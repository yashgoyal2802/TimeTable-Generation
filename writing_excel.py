from openpyxl import load_workbook

# faculty_list=[[[0, 1, 2, 3], [1, 0, 3, 2], [2, 3, 0, 1], [3, 2, 1, -1], [0, 1, 2, 3]],
#               [[1, 0, 3, 2], [2, 3, 0, 1], [3, 2, 1, -1], [0, 1, 2, 3], [1, 0, 3, 2]],
#               [[2, 3, 0, 1], [3, 2, 0, 1], [0, 1, 2, 3], [1, 0, 3, 2], [2, 3, -1, 1]],
#               [[3, 2, -1, 1], [1, 0, 2, -1], [2, 3, -1, 1], [2, 3, -1, -1], [-1, -1, 2, -1]],
#               [[-1, -1, 2, -1], [-1, -1, 2, -1], [-1, -1, -1, 2], [-1, -1, -1, 2], [-1, -1, -1, -1]],
#               [[-1, -1, -1, -1], [-1, -1, -1, -1], [-1, -1, -1, -1], [-1, -1, -1, -1], [-1, -1, -1, -1]],
#               [[-1, -1, -1, -1], [-1, -1, -1, -1], [-1, -1, -1, -1], [-1, -1, -1, -1], [-1, -1, -1, -1]],
#               [[-1, -1, -1, -1], [-1, -1, -1, -1], [-1, -1, -1, -1], [-1, -1, -1, -1], [-1, -1, -1, -1]]]


def write(faculty_list):
    try:
        branch = {0: 'B Tech CS', 1: 'B Tech IT', 2: 'MBA Tech CS', 3: 'MBA Tech IT', -1: ''}

        book = load_workbook('Timetable.xlsx')
        sheet = book["Sheet1"]

    #sheet.cell(row=4, column=2).value = 1

        r, c = 4, 2
        for hour in faculty_list:
            for day in hour:
                for fac_no in day:
                    sheet.cell(row=r, column=c).value = branch[fac_no]
                    c = c+1
            if r == 7:
                r = r+2
            else:
                r = r+1
            c = 2

        book.save('Timetable.xlsx')
    except IOError:
        print("An Error occurred trying to write in the file.")